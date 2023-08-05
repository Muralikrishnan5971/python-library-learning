from openpyxl import workbook, load_workbook

# load our workbook
wb = load_workbook("sample3.xlsx")  # openpyxl module supports only xlsx format

sheets = wb.sheetnames      # returns a list of sheet names
print(sheets)

active_sheet = wb.active        # returns the name of the active sheet
print(active_sheet)

print(active_sheet['A1'].value)
print(active_sheet['A2'].value)
print(active_sheet['A3'].value)

active_sheet['A2'] = 1
active_sheet['A3'] = 2

print(active_sheet['B1'].value)
print(active_sheet['B2'].value)
# print(active_sheet['B3'].value)

active_sheet['B3'].value = "ORange"      # to write to a cell
print(active_sheet['B3'].value)

print(wb['MyLinks'])        # To access a sheet